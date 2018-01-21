import openpyxl
from matplotlib import pyplot as plt
import matplotlib.style as style
import matplotlib
import seaborn as sns; sns.set()
import operator
import warnings; warnings.simplefilter("ignore")
style.use('fivethirtyeight')


wb = openpyxl.load_workbook('EO.xlsx')
sheet = wb.get_active_sheet()
print("This sheet is named {}.".format(sheet.title))
enrl = str(sheet.cell(row = 1, column = 14).value)
prdct_type = str(sheet.cell(row = 1, column = 9).value)


agnt_mdcl_set = set()
agnt_anc_set = set()


for cntr in range(1, 50000):
    litem = str(sheet.cell(row = cntr + 1, column = 14).value)
    prdct = str(sheet.cell(row = cntr + 1, column = 9).value)
    if prdct == 'MG' or prdct == 'MA':
        agnt_mdcl_set.add(litem)
    if prdct == 'D' or prdct == 'V':
        agnt_anc_set.add(litem)


agnt_mdcl_list = sorted(list(agnt_mdcl_set))
agnt_anc_list = sorted(list(agnt_anc_set))


mdcl_dctnry = {}
anc_dctnry = {}


for i in agnt_mdcl_list:
    mdcl_dctnry[i] = 0

for i in agnt_anc_list:
    anc_dctnry[i] = 0


for cntr in range(50000):
    prdct = str(sheet.cell(row = cntr + 1, column = 9).value)
    agnt = str(sheet.cell(row = cntr + 1, column = 14).value)
    if prdct == 'MG' or prdct == 'MA':
        mdcl_dctnry[agnt] += 1
    if prdct == 'D' or prdct == 'V':
        anc_dctnry[agnt] += 1


sortedlistmdc = sorted(mdcl_dctnry.items(), key=operator.itemgetter(1))
sortedlistanc = sorted(anc_dctnry.items(), key=operator.itemgetter(1))


agntlistmdc = []
slslistmdc = []
for a, b in sortedlistmdc:
    agntlistmdc.append(a)
    slslistmdc.append(b)


agntlistanc = []
slslistanc = []
for a, b in sortedlistanc:
    agntlistanc.append(a)
    slslistanc.append(b)


matplotlib.rcParams.update({'ytick.labelsize': 11})
plt.figure(figsize=(12, 14))


ax = plt.subplot(211)
plt.barh(range(len(agntlistmdc)), slslistmdc, edgecolor = 'black', color = '#37ce55', alpha = 0.8, linewidth = 0.5)
ax.set_yticks(range(len(agntlistmdc)))
ax.set_yticklabels(agntlistmdc)
ax.grid(color='#399aef', alpha = 0.15)
ax.legend(["MDC Sales"])


ax2 = plt.subplot(212)
plt.barh(range(len(agntlistanc)), slslistanc, edgecolor = 'black', color = '#f2d33a', alpha = 0.8, linewidth = 0.5)
ax2.set_yticks(range(len(agntlistanc)))
ax2.set_yticklabels(agntlistanc)
ax2.grid(color='#ef3939', alpha = 0.15)
ax2.legend(["ANC Sales"])


plt.title('Sales Month to Date')
plt.tight_layout()


plt.savefig('Sales.png')


plt.show()
