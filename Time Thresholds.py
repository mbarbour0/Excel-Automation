import operator
import warnings; warnings.simplefilter("ignore")

import matplotlib
from matplotlib import pyplot as plt
import matplotlib.style as style
import openpyxl
import seaborn as sns; sns.set()

style.use('fivethirtyeight')

wb2 = openpyxl.load_workbook('Book1.xlsx')
sheet2 = wb2.get_sheet_by_name('Sheet_1')
print("This sheet is titled {}.".format(sheet2.title))


def get_sec(time_str):
    """Returns time in seconds"""
    try:
        h, m, s = time_str.split(':')
        return int(h) * 3600 + int(m) * 60 + int(s)
    except ValueError:
        pass


def brkl(arg1):
    """Calculates estimated break and lunch times"""
    if arg1 < 4:
        return 0
    elif arg1 < 6:
        return 0.35
    elif arg1 < 8:
        return 0.85
    elif arg1 < 12:
        return 1.1
    elif arg1 < 14:
        return 1.45
    elif arg1 < 16:
        return 1.95
    elif arg1 < 20:
        return 2.2
    elif arg1 < 22:
        return 2.55
    elif arg1 < 24:
        return 3.05
    elif arg1 < 28:
        return 3.3
    elif arg1 < 30:
        return 3.65
    elif arg1 < 32:
        return 4.15


"""Locations of source data"""
# avlb = str(sheet.cell(row=9, column=16).value)
# ibnd = str(sheet.cell(row=9, column=17).value)
# outb = str(sheet.cell(row=9, column=18).value)
# hold = str(sheet.cell(row=9, column=20).value)
# ttl = str(sheet.cell(row=9, column=21).value)

dict1 = {}

for i in range(1, 75):
    if sheet2.cell(row = i + 9, column = 4).value != None:
        name = str(sheet2.cell(row = i + 9, column = 3).value)
        a = str(sheet2.cell(row = i + 9, column = 16).value)
        b = str(sheet2.cell(row = i + 9, column = 17).value)
        c = str(sheet2.cell(row = i + 9, column = 18).value)
        d = str(sheet2.cell(row = i + 9, column = 20).value)
        e = str(sheet2.cell(row = i + 9, column = 21).value)
        aout = get_sec(a)
        bout = get_sec(b)
        cout = get_sec(c)
        dout = get_sec(d)
        eout = get_sec(e)
        try:
            result = round(((aout + bout + cout + dout) / (eout - (brkl(eout/3600)*3600))) * 100, 2)
        except TypeError:
            continue
        dict1[name] = result

dict2 =  {}

for k, v, in dict1.items():
    if v > 50:
        dict2[k] = v

sorted_dict2 = sorted(dict2.items(), key=operator.itemgetter(1))

names = []
results = []

for a, b in sorted_dict2:
    names.append(a)
    results.append(b)

"""Plotting Efficiency of Time Used"""

matplotlib.rcParams.update({'xtick.labelsize': 14})
plt.figure(figsize=(22, 10))
plt.bar(range(len(names)), results, edgecolor = 'black', color = '#39c7f1', alpha = 0.7, linewidth = 3.0)

ax = plt.subplot()
ax.set_xticks(range(len(names)))
ax.set_xticklabels(names)
ax.grid(color='red')

plt.ylabel('Efficiency')
plt.xlabel('Colleague')
plt.title('Weekly Efficiency 1-16-2018 to 1-18-2018')

plt.axis(ymin = 50, ymax = 100)

for tick in ax.get_xticklabels():
    tick.set_rotation(90)

plt.tight_layout()

plt.savefig('Efficiency.png')

plt.show()
